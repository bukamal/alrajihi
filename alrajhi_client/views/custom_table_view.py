# -*- coding: utf-8 -*-
from PyQt5.QtWidgets import QTableView, QMenu, QAction, QFileDialog, QMessageBox, QApplication, QHeaderView, QStyledItemDelegate
from PyQt5.QtCore import Qt, QSettings, QTimer
from PyQt5.QtGui import QKeySequence
from ui.table_keyboard_policy import StandardTableKeyboardMixin
from theme_manager import ThemeManager
from views.widgets.components.table_preferences import TablePreferences
from i18n import translate
import re

class CenterAlignDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        option.displayAlignment = Qt.AlignCenter
        super().paint(painter, option, index)

class CustomTableView(StandardTableKeyboardMixin, QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableView.SelectRows)
        self.setSelectionMode(QTableView.ExtendedSelection)
        self.setSortingEnabled(True)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_menu)

        self._restoring_layout = False
        self._layout_save_pending = False
        self._preferences = TablePreferences()
        self._column_contract = None
        self.horizontalHeader().setSectionsMovable(True)
        self.horizontalHeader().setStretchLastSection(False)
        self.horizontalHeader().sectionResized.connect(self._schedule_save_layout)
        self.horizontalHeader().sectionMoved.connect(lambda *args: self._schedule_save_layout())

        self.setItemDelegate(CenterAlignDelegate(self))

        self.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.verticalHeader().setDefaultAlignment(Qt.AlignCenter)

        self.copy_action = QAction(self)
        self.copy_action.setShortcut(QKeySequence.Copy)
        self.copy_action.triggered.connect(self.copy_selection)
        self.addAction(self.copy_action)
        self.init_standard_table_keyboard()

        self.refresh_style()

    def set_table_identity(self, identity: str):
        """Set a stable key used to persist column widths/order/visibility."""
        if identity:
            self.setObjectName(identity)
            self.restore_layout()

    def set_column_contract(self, page_id: str = "", table_id: str = "", contract_id: str = "") -> None:
        """Attach a universal column contract to this table.

        Display, print and export preferences are read from the same settings
        keys used by the settings surface.  The table applies display
        visibility immediately when a model is available.
        """
        try:
            from workspace.tables import table_column_contract, table_column_contract_by_id
            contract = table_column_contract_by_id(contract_id) if contract_id else table_column_contract(page_id, table_id)
        except Exception:
            contract = None
        self._column_contract = contract
        if contract is not None:
            self.setProperty("column_contract_id", contract.contract_id)
            if not self.objectName():
                self.setObjectName(contract.contract_id.replace(".", "_"))
            self._apply_contract_display_visibility()

    def column_contract(self):
        return self._column_contract

    def _apply_contract_display_visibility(self) -> None:
        model = self._source_model_for_columns()
        table_model = self.model()
        contract = self._column_contract
        if contract is None or model is None or table_model is None:
            return
        try:
            from workspace.settings.column_preferences import display_keys_for_contract
            allowed = set(display_keys_for_contract(contract))
        except Exception:
            allowed = set()
        if not allowed:
            return
        model_keys = [self._column_key_for_index(col) for col in range(table_model.columnCount())]
        # If the current model cannot be mapped to contract keys, leave it as-is
        # rather than hiding columns accidentally.
        if not (set(model_keys) & allowed):
            return
        visible_count = 0
        for col, key in enumerate(model_keys):
            visible = key in allowed
            self.setColumnHidden(col, not visible)
            if visible:
                visible_count += 1
        if visible_count == 0 and table_model.columnCount() > 0:
            self.setColumnHidden(0, False)

    def show_contract_column_customizer(self) -> bool:
        contract = self._column_contract
        if contract is None:
            return False
        try:
            from views.dialogs.column_contract_customizer import ColumnContractCustomizerDialog
            dialog = ColumnContractCustomizerDialog(self, contract.contract_id)
            if dialog.exec_():
                self._apply_contract_display_visibility()
                self.save_layout()
                return True
        except Exception as exc:
            QMessageBox.warning(self, translate("warning"), str(exc))
        return False

    def _source_model_for_columns(self):
        return getattr(self, "_source_model", None) or self.model()

    def _column_key_for_index(self, column: int) -> str:
        model = self._source_model_for_columns()
        if model is None:
            return str(column)
        columns = getattr(model, "columns", None)
        if columns is not None and 0 <= column < len(columns):
            try:
                return str(columns[column].key)
            except Exception:
                pass
        data_keys = getattr(model, "_data_keys", None)
        if data_keys is not None and 0 <= column < len(data_keys):
            return str(data_keys[column])
        return str(column)

    def _columns_for_purpose(self, purpose: str = "display") -> list[int]:
        model = self.model()
        if not model:
            return []
        purpose = str(purpose or "display").strip().lower()
        display_cols = [col for col in range(model.columnCount()) if not self.isColumnHidden(col)]
        contract = self._column_contract
        if contract is None:
            return display_cols
        try:
            from workspace.tables import keys_for_output
            normalized = "print" if purpose in {"print", "printing"} else ("export" if purpose in {"export", "excel", "xlsx", "csv"} else "display")
            allowed = set(keys_for_output(contract, normalized))
            if not allowed:
                return []
            # Display follows the user's visible/hidden state.  Printing/exporting
            # use their own contract/settings flags so a user can hide an on-screen
            # column without removing it from official output.
            candidate_cols = display_cols if normalized == "display" else list(range(model.columnCount()))
            return [col for col in candidate_cols if self._column_key_for_index(col) in allowed]
        except Exception:
            return display_cols

    def _layout_key(self) -> str:
        name = self.objectName() or self.__class__.__name__
        parent = self.parent()
        parent_name = parent.__class__.__name__ if parent is not None else 'root'
        return f"tables/{parent_name}/{name}"

    def _schedule_save_layout(self, *args):
        if self._restoring_layout:
            return
        if self._layout_save_pending:
            return
        self._layout_save_pending = True
        QTimer.singleShot(250, self.save_layout)

    def save_layout(self):
        self._layout_save_pending = False
        if not self.model():
            return
        self._preferences.save_state(self._layout_key(), self.horizontalHeader().saveState())

    def restore_layout(self):
        if not self.model():
            return
        state = self._preferences.load_state(self._layout_key())
        if state:
            self._restoring_layout = True
            try:
                self.horizontalHeader().restoreState(state)
            finally:
                self._restoring_layout = False

    def reset_layout(self):
        self._preferences.reset(self._layout_key())
        self._restoring_layout = True
        try:
            for col in range(self.model().columnCount() if self.model() else 0):
                self.setColumnHidden(col, False)
            self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        finally:
            self._restoring_layout = False
        self.save_layout()

    def setModel(self, model):
        super().setModel(model)
        self.restore_layout()
        self._apply_contract_display_visibility()

    def copy_selection(self):
        selection = self.selectionModel().selectedIndexes()
        if not selection:
            return
        rows = sorted(set(i.row() for i in selection))
        cols = sorted(set(i.column() for i in selection))
        text = ""
        for r in rows:
            row_data = []
            for c in cols:
                idx = self.model().index(r, c)
                data = self.model().data(idx, Qt.DisplayRole)
                row_data.append(str(data))
            text += "\t".join(row_data) + "\n"
        QApplication.clipboard().setText(text)

    def _show_menu(self, pos):
        menu = QMenu()
        export_excel = QAction(translate("excel"), self)
        export_excel.triggered.connect(self.export_to_excel)
        menu.addAction(export_excel)
        print_action = QAction(translate("print"), self)
        print_action.triggered.connect(self.print_table)
        menu.addAction(print_action)
        menu.addSeparator()
        copy_act = QAction(translate("copy"), self)
        copy_act.triggered.connect(self.copy_selection)
        menu.addAction(copy_act)

        model = self.model()
        if model:
            columns_menu = menu.addMenu(translate("columns"))
            for col in range(model.columnCount()):
                header = model.headerData(col, Qt.Horizontal, Qt.DisplayRole) or translate("column_number", number=col + 1)
                action = QAction(str(header), self)
                action.setCheckable(True)
                action.setChecked(not self.isColumnHidden(col))
                action.toggled.connect(lambda checked, c=col: self.set_column_visible(c, checked))
                columns_menu.addAction(action)
            if self._column_contract is not None:
                contract_columns = QAction(translate("settings_surface_column_editor"), self)
                contract_columns.triggered.connect(self.show_contract_column_customizer)
                columns_menu.addSeparator()
                columns_menu.addAction(contract_columns)
            reset_columns = QAction(translate("reset_columns"), self)
            reset_columns.triggered.connect(self.reset_layout)
            columns_menu.addSeparator()
            columns_menu.addAction(reset_columns)
        menu.exec(self.viewport().mapToGlobal(pos))

    def set_column_visible(self, column, visible):
        self.setColumnHidden(column, not visible)
        self.save_layout()

    def _set_column_visible(self, column, visible):
        self.set_column_visible(column, visible)

    def export_to_excel(self):
        model = self.model()
        if not model:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, translate("warning"), translate("openpyxl_missing"))
            return
        filename, _ = QFileDialog.getSaveFileName(self, translate("save_report"), "report.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = translate("report")
        visible_cols = self._columns_for_purpose('export')
        for out_col, col in enumerate(visible_cols, start=1):
            header = model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
            cell = ws.cell(row=1, column=out_col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in range(model.rowCount()):
            for out_col, col in enumerate(visible_cols, start=1):
                idx = model.index(row, col)
                value = model.data(idx, Qt.DisplayRole)
                ws.cell(row=row+2, column=out_col, value=value)
        wb.save(filename)
        QMessageBox.information(self, "نجاح", f"تم التصدير إلى {filename}")

    def print_table(self, mode='preview'):
        """Print table through the centralized printing service.

        Supports any Qt model already used by project pages, respects hidden columns,
        and uses the unified company/report template instead of a local ad-hoc HTML.
        """
        model = self.model()
        if not model:
            return
        visible_cols = self._columns_for_purpose('print')
        headers = []
        for col in visible_cols:
            h = model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
            headers.append(str(h) if h else f"عمود {col + 1}")

        rows = []
        for row in range(model.rowCount()):
            row_data = []
            for col in visible_cols:
                idx = model.index(row, col)
                val = model.data(idx, Qt.DisplayRole)
                row_data.append(str(val) if val is not None else '')
            rows.append(row_data)

        title = self.property('print_title') or self.windowTitle() or self.objectName() or 'تقرير جدول'
        subtitle = f"عدد السجلات: {len(rows)}"
        try:
            from printing.printing_service import printing_service
            if mode == 'browser':
                printing_service.report_print(str(title), rows, headers, self, subtitle=subtitle)
            elif mode == 'direct':
                printing_service.report_print(str(title), rows, headers, self, subtitle=subtitle)
            elif mode == 'pdf':
                # Phase 235: legacy PDF mode follows unified print output.
                printing_service.report_print(str(title), rows, headers, self, subtitle=subtitle)
            else:
                printing_service.report_print(str(title), rows, headers, self, subtitle=subtitle)
        except Exception as exc:
            QMessageBox.warning(self, "تعذر الطباعة", str(exc))

    def refresh_style(self):
        self.setStyleSheet(ThemeManager.get_stylesheet())
        self.viewport().update()

    def showEvent(self, event):
        self.refresh_style()
        super().showEvent(event)


